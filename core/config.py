import json
import os
import winreg

REG_PATH = r"Software\DailyTaskLog"

DEFAULTS = {
    "reminder_enabled": True,
    "reminder_time": "17:30",
    "theme": "light",
    "excel_path": os.path.join(os.path.expanduser("~"), "Documents", "task_log.xlsx"),
    "projects": {"General": [], "Admin": []},
    "autostart": False,
}


def _read_excel_config(path: str) -> dict:
    """Read key/value pairs from the Config sheet of the given Excel file."""
    from openpyxl import load_workbook
    result = {}
    try:
        wb = load_workbook(path, read_only=True)
        if "Config" in wb.sheetnames:
            ws = wb["Config"]
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                if not row or not row[0]:
                    continue
                key, val = row[0], row[1]
                if val is None:
                    continue
                if val == "True":
                    val = True
                elif val == "False":
                    val = False
                else:
                    try:
                        val = json.loads(val)
                    except Exception:
                        pass
                result[key] = val
        wb.close()
    except Exception:
        pass
    return result


def load():
    cfg = dict(DEFAULTS)

    try:
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, REG_PATH, 0, winreg.KEY_READ) as key:
            excel_path, _ = winreg.QueryValueEx(key, "excel_path")
            cfg["excel_path"] = excel_path
            try:
                autostart, _ = winreg.QueryValueEx(key, "autostart")
                cfg["autostart"] = bool(autostart)
            except Exception:
                pass
    except Exception:
        pass

    if isinstance(cfg.get("projects"), list):
        cfg["projects"] = {p: [] for p in cfg["projects"]}

    excel_path = cfg.get("excel_path")
    if excel_path and os.path.exists(excel_path):
        cfg.update(_read_excel_config(excel_path))

    return cfg


def save(cfg):
    excel_path = cfg.get("excel_path")
    autostart = cfg.get("autostart", False)
    
    if excel_path:
        try:
            with winreg.CreateKey(winreg.HKEY_CURRENT_USER, REG_PATH) as key:
                winreg.SetValueEx(key, "excel_path", 0, winreg.REG_SZ, excel_path)
                winreg.SetValueEx(key, "autostart", 0, winreg.REG_DWORD, int(autostart))
        except Exception:
            pass

    if excel_path:
        from openpyxl import load_workbook, Workbook
        try:
            os.makedirs(os.path.dirname(excel_path), exist_ok=True)
            if os.path.exists(excel_path):
                wb = load_workbook(excel_path)
            else:
                wb = Workbook()
                wb.active.title = "Task Log"
            
            if "Config" in wb.sheetnames:
                ws = wb["Config"]
                ws.delete_rows(1, ws.max_row + 1)
            else:
                ws = wb.create_sheet("Config")
            
            ws.cell(row=1, column=1, value="Setting")
            ws.cell(row=1, column=2, value="Value")
            
            row = 2
            for k, v in cfg.items():
                if k == "excel_path":
                    continue
                if isinstance(v, (dict, list)):
                    val_str = json.dumps(v)
                else:
                    val_str = str(v)
                ws.cell(row=row, column=1, value=k)
                ws.cell(row=row, column=2, value=val_str)
                row += 1
            wb.save(excel_path)
            wb.close()
        except Exception:
            pass


def set_autostart(enabled: bool):
    import sys
    key_path = r"Software\Microsoft\Windows\CurrentVersion\Run"
    app_name = "DailyTaskLog"
    exe_path = sys.executable if getattr(sys, "frozen", False) else None
    if exe_path is None:
        return
    try:
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, key_path, 0, winreg.KEY_SET_VALUE)
        if enabled:
            winreg.SetValueEx(key, app_name, 0, winreg.REG_SZ, f'"{exe_path}"')
        else:
            try:
                winreg.DeleteValue(key, app_name)
            except FileNotFoundError:
                pass
        winreg.CloseKey(key)
    except Exception:
        pass
