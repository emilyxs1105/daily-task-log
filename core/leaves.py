import os
from datetime import date, datetime

TYPES = {
    "public_holiday": "Public Holiday",
    "annual_leave":   "Annual Leave",
    "mc_leave":       "MC Leave",
}

ICONS = {
    "public_holiday": "🏖",
    "annual_leave":   "🌴",
    "mc_leave":       "🤒",
}


def _get_excel_path() -> str | None:
    from core import config
    return config.load().get("excel_path")


def _read_leaves_from_ws(ws) -> dict:
    data = {}
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if not row or not row[0]:
            continue
        d_val, type_val = row[0], row[1]
        if d_val is None or type_val is None:
            continue
        if isinstance(d_val, (date, datetime)):
            d_str = d_val.strftime("%Y-%m-%d")
        else:
            d_str = str(d_val).strip()
        data[d_str] = type_val
    return data


def _write_leaves_to_ws(ws, data: dict):
    ws.delete_rows(1, ws.max_row + 1)
    ws.cell(row=1, column=1, value="Date")
    ws.cell(row=1, column=2, value="Type")
    for i, (k, v) in enumerate(sorted(data.items()), 2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)


def load(excel_path: str = None) -> dict:
    if excel_path is None:
        excel_path = _get_excel_path()
    if not excel_path or not os.path.exists(excel_path):
        return {}
    from openpyxl import load_workbook
    try:
        wb = load_workbook(excel_path, read_only=True)
        if "Leaves" in wb.sheetnames:
            data = _read_leaves_from_ws(wb["Leaves"])
        else:
            data = {}
        wb.close()
        return data
    except Exception:
        return {}


def _mutate(excel_path: str, fn):
    """Open workbook once, read Leaves data, apply fn, write back — single round-trip."""
    from openpyxl import load_workbook, Workbook
    try:
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
        else:
            wb = Workbook()
            wb.active.title = "Task Log"
        if "Leaves" in wb.sheetnames:
            ws = wb["Leaves"]
            data = _read_leaves_from_ws(ws)
        else:
            ws = wb.create_sheet("Leaves")
            data = {}
        fn(data)
        _write_leaves_to_ws(ws, data)
        wb.save(excel_path)
        wb.close()
    except Exception:
        pass


def mark(d: date, leave_type: str):
    excel_path = _get_excel_path()
    if excel_path:
        _mutate(excel_path, lambda data: data.update({d.strftime("%Y-%m-%d"): leave_type}))


def clear(d: date):
    excel_path = _get_excel_path()
    if excel_path:
        _mutate(excel_path, lambda data: data.pop(d.strftime("%Y-%m-%d"), None))


def get(d: date) -> str | None:
    excel_path = _get_excel_path()
    return load(excel_path).get(d.strftime("%Y-%m-%d"))
