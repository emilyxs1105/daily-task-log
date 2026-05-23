import os
from datetime import date, datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADERS = ["ID", "Date", "Project", "Task", "Hours", "Notes"]
DATE_FMT = "%Y-%m-%d"
_SEP = " > "
_STATUS_SHEET = "Project Status"
_TODO_SHEET   = "To-Do"


def _parse_project(raw: str) -> tuple[str, str]:
    if raw and _SEP in raw:
        proj, sub = raw.split(_SEP, 1)
        return proj.strip(), sub.strip()
    return (raw or ""), ""


def _fmt_project(project: str, subproject: str) -> str:
    return f"{project}{_SEP}{subproject}" if subproject else project


def _ensure_file(path: str):
    if not os.path.exists(path):
        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Task Log"
        _write_headers(ws)
        wb.save(path)
    return path


def _write_headers(ws):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="185FA5")
    border = Border(bottom=Side(style="thin", color="185FA5"))
    widths = [8, 12, 18, 40, 8, 30]
    for col, (h, w) in enumerate(zip(HEADERS, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _next_id(ws) -> int:
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and isinstance(row[0], int):
            max_id = max(max_id, row[0])
    return max_id + 1


def load_all(path: str) -> list[dict]:
    _ensure_file(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rid, d, project, task, hours, notes = (list(row) + [None] * 6)[:6]
        if d is None:
            continue
        if isinstance(d, datetime):
            d = d.date()
        elif isinstance(d, str):
            try:
                d = datetime.strptime(d, DATE_FMT).date()
            except Exception:
                continue
        proj, sub = _parse_project(project or "")
        entries.append({
            "id": rid,
            "date": d,
            "project": proj,
            "subproject": sub,
            "task": task or "",
            "hours": float(hours) if hours else 0.0,
            "notes": notes or "",
        })
    wb.close()
    entries.sort(key=lambda x: x["date"], reverse=True)
    return entries


def save_entry(path: str, entry: dict) -> int:
    _ensure_file(path)
    wb = load_workbook(path)
    ws = wb.active
    if not ws["A1"].value or ws["A1"].value != "ID":
        _write_headers(ws)
    new_id = _next_id(ws)
    d = entry["date"]
    if isinstance(d, date):
        d = d.strftime(DATE_FMT)
    ws.append([new_id, d, _fmt_project(entry["project"], entry.get("subproject", "")),
               entry["task"], entry["hours"], entry.get("notes", "")])
    _style_data_row(ws, ws.max_row)
    wb.save(path)
    wb.close()
    return new_id


def update_entry(path: str, entry: dict):
    _ensure_file(path)
    wb = load_workbook(path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == entry["id"]:
            d = entry["date"]
            if isinstance(d, date):
                d = d.strftime(DATE_FMT)
            row[0].value = entry["id"]
            row[1].value = d
            row[2].value = _fmt_project(entry["project"], entry.get("subproject", ""))
            row[3].value = entry["task"]
            row[4].value = entry["hours"]
            row[5].value = entry.get("notes", "")
            break
    wb.save(path)
    wb.close()


def delete_entry(path: str, entry_id: int):
    _ensure_file(path)
    wb = load_workbook(path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == entry_id:
            ws.delete_rows(row[0].row)
            break
    wb.save(path)
    wb.close()


def _style_data_row(ws, row_num: int):
    border = Border(bottom=Side(style="thin", color="E8E8E8"))
    fill_even = PatternFill("solid", fgColor="F7FAFD")
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.border = border
        if row_num % 2 == 0:
            cell.fill = fill_even
        cell.alignment = Alignment(vertical="center", wrap_text=(col == 4))


def export_filtered(path: str, entries: list[dict], out_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Exported Log"
    _write_headers(ws)
    for entry in entries:
        d = entry["date"]
        if isinstance(d, date):
            d = d.strftime(DATE_FMT)
        ws.append([entry["id"], d, _fmt_project(entry["project"], entry.get("subproject", "")),
                   entry["task"], entry["hours"], entry.get("notes", "")])
        _style_data_row(ws, ws.max_row)
    wb.save(out_path)
    wb.close()


def get_missed_days(entries: list[dict], days_back: int = 30,
                    leaves: dict = None) -> list[date]:
    today = date.today()
    logged = {e["date"] for e in entries}
    leave_dates = set(leaves or {})
    missed = []
    for i in range(1, days_back + 1):
        day = today - timedelta(days=i)
        if (day.weekday() < 5
                and day not in logged
                and day.strftime("%Y-%m-%d") not in leave_dates):
            missed.append(day)
    return sorted(missed, reverse=True)


def get_last_day_entries(entries: list[dict]) -> list[dict]:
    if not entries:
        return []
    latest = entries[0]["date"]
    return [e for e in entries if e["date"] == latest]


def load_project_statuses(path: str) -> dict:
    _ensure_file(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    if _STATUS_SHEET not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb[_STATUS_SHEET]
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            result[str(row[0])] = str(row[1]) if row[1] else ""
    wb.close()
    return result


def save_project_status(path: str, project: str, status: str):
    _ensure_file(path)
    wb = load_workbook(path)
    if _STATUS_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(_STATUS_SHEET)
        ws.append(["Project", "Status"])
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill("solid", fgColor="185FA5")
        for col in range(1, 3):
            cell = ws.cell(row=1, column=col)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 18
    else:
        ws = wb[_STATUS_SHEET]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == project:
            row[1].value = status
            wb.save(path)
            wb.close()
            return
    ws.append([project, status])
    wb.save(path)
    wb.close()


# ── To-Do ──────────────────────────────────────────────────────────────────────

def _ensure_todo_sheet(wb):
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="185FA5")
    cols = ["ID", "Task", "Done", "Created", "Project", "Subproject"]
    widths = [8, 50, 8, 14, 20, 20]
    if _TODO_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(_TODO_SHEET)
        ws.append(cols)
        for col, w in zip(range(1, len(cols) + 1), widths):
            cell = ws.cell(row=1, column=col)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = w
    else:
        ws = wb[_TODO_SHEET]
        # Migrate: add Project/Subproject headers if this is an older sheet
        for col, (header, w) in enumerate(zip(cols[4:], widths[4:]), 5):
            if ws.cell(row=1, column=col).value != header:
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(col)].width = w
    return ws


def load_todos(path: str) -> list[dict]:
    _ensure_file(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    todos = []
    if _TODO_SHEET in wb.sheetnames:
        ws = wb[_TODO_SHEET]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            cols = list(row) + [None] * 6
            tid, task, done, created, project, subproject = cols[:6]
            if task is None:
                continue
            todos.append({
                "id":         tid,
                "task":       str(task),
                "done":       bool(done),
                "created":    str(created) if created else "",
                "project":    str(project) if project else "",
                "subproject": str(subproject) if subproject else "",
            })
    wb.close()
    return todos


def add_todo(path: str, task: str) -> int:
    _ensure_file(path)
    wb = load_workbook(path)
    ws = _ensure_todo_sheet(wb)
    max_id = max((row[0].value for row in ws.iter_rows(min_row=2)
                  if row[0].value and isinstance(row[0].value, int)), default=0)
    new_id = max_id + 1
    ws.append([new_id, task, False, date.today().strftime(DATE_FMT), None, None])
    wb.save(path)
    wb.close()
    return new_id


def update_todo_done(path: str, todo_id: int, done: bool):
    _ensure_file(path)
    wb = load_workbook(path)
    if _TODO_SHEET in wb.sheetnames:
        for row in wb[_TODO_SHEET].iter_rows(min_row=2):
            if row[0].value == todo_id:
                row[2].value = done
                break
    wb.save(path)
    wb.close()


def update_todo_project(path: str, todo_id: int, project: str, subproject: str):
    _ensure_file(path)
    wb = load_workbook(path)
    ws = _ensure_todo_sheet(wb)
    for row in ws.iter_rows(min_row=2):
        if row[0].value == todo_id:
            ws.cell(row=row[0].row, column=5, value=project or None)
            ws.cell(row=row[0].row, column=6, value=subproject or None)
            break
    wb.save(path)
    wb.close()


def delete_todo(path: str, todo_id: int):
    _ensure_file(path)
    wb = load_workbook(path)
    if _TODO_SHEET in wb.sheetnames:
        ws = wb[_TODO_SHEET]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == todo_id:
                ws.delete_rows(row[0].row)
                break
    wb.save(path)
    wb.close()
